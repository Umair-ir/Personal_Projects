import pandas as pd
import calendar
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Alignment
import traceback

def generate_sheet_names(month_number):
    year = 2024  # You can use any specific year or the current year using datetime
    days_in_month = calendar.monthrange(year, month_number)[1]
    # Use '_' instead of '/' to ensure valid sheet names
    sheet_names = [f'{month_number}-{str(day).zfill(2)}' for day in range(1, days_in_month + 1)]
    return sheet_names

def link_sheets(wb, sheet_names):
    # Start linking from the second sheet
    for i in range(1, len(sheet_names)):
        current_sheet = wb[sheet_names[i]]
        previous_sheet = wb[sheet_names[i - 1]]

        # Example of copying data from C3:C4 into B3:B4 in the next sheet
        current_sheet['B3'] = f"='{previous_sheet.title}'!C3"
        current_sheet['B4'] = f"='{previous_sheet.title}'!C4"

        for row in range(9, 25):
            current_sheet[f'B{row}'] = f"='{previous_sheet.title}'!C{row}"

        # Copy C29:C37 into B29:B37
        for row in range(29, 38):
            current_sheet[f'B{row}'] = f"='{previous_sheet.title}'!C{row}"

        for row in range(42, 56):
            current_sheet[f'B{row}'] = f"='{previous_sheet.title}'!C{row}"
        
        for row in range(3, 21):
            current_sheet[f'G{row}'] = f"='{previous_sheet.title}'!H{row}"
        
        for row in range(25, 31):
            current_sheet[f'G{row}'] = f"='{previous_sheet.title}'!H{row}"

        for row in range(35, 38):
            current_sheet[f'G{row}'] = f"='{previous_sheet.title}'!H{row}"

        for row in range(42, 47):
            current_sheet[f'G{row}'] = f"='{previous_sheet.title}'!H{row}"

        for row in range(51, 53):
            current_sheet[f'G{row}'] = f"='{previous_sheet.title}'!H{row}"

sheetDate = input("Enter the number of the month (1-12): ")
print(f"Input received: {sheetDate}")

try:
    month_number = int(sheetDate)
    print(f"Converted month number: {month_number}")

    if 1 <= month_number <= 12:
        sheet_names = generate_sheet_names(month_number)

        wb = Workbook()

        green_fill = PatternFill(start_color="71C562", end_color="71C562", fill_type="solid")
        blue_fill = PatternFill(start_color="90E0EF", end_color="90E0EF", fill_type="solid")
        LightBlue_fill = PatternFill(start_color="CAF0F8", end_color="CAF0F8", fill_type="solid")
        lavender_fill = PatternFill(start_color="e1affd", end_color="E6E6FA", fill_type="solid")
        currency_style = NamedStyle(name="currency_style", number_format='$#,##0.00')

        #Medium Border style
        medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Define a thin border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Function to apply border to a range of cells
        def apply_border(ws, cell_range):
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = thin_border

        # Create sheets and add data with formulas and formatting
        for sheet_name in sheet_names:
            ws = wb.create_sheet(title=sheet_name)


            ws.merge_cells('K21:M21')
            ws.merge_cells('K2:M2')
            ws.merge_cells('L24:M24')
            ws.merge_cells('L26:M26')

            ws['K2'] = 'TOTALS'

            #Dollar Amounts
            ws['A2'] = 1
            ws['A2'].style = currency_style 
            ws['A8'] = 2
            ws['A8'].style = currency_style
            ws['A28'] = 3
            ws['A28'].style = currency_style
            ws['A41'] = 5
            ws['A41'].style = currency_style
            ws['F2'] = 10
            ws['F2'].style = currency_style
            ws['F24'] = 20
            ws['F24'].style = currency_style
            ws['F34'] = 25
            ws['F34'].style = currency_style
            ws['F41'] = 30
            ws['F41'].style = currency_style
            ws['F50'] = 50
            ws['F50'].style = currency_style

            ws['K4'] = 1
            ws['K4'].style = currency_style

            ws['K5'] = 2
            ws['K5'].style = currency_style

            ws['K6'] = 3
            ws['K6'].style = currency_style

            ws['K7'] = 5
            ws['K7'].style = currency_style

            ws['K8'] = 10
            ws['K8'].style = currency_style

            ws['K9'] = 20
            ws['K9'].style = currency_style

            ws['K10'] = 25
            ws['K10'].style = currency_style

            ws['K11'] = 30
            ws['K11'].style = currency_style

            ws['K12'] = 50
            ws['K12'].style = currency_style

            
            #Totals
            ws['A5'] = 'TOTAL'
            ws['B5'] = '==SUM(B3:B4)'
            ws['A25'] = 'TOTAL'
            ws['A38'] = 'TOTAL'
            ws['A57'] = 'TOTAL'
            ws['F21'] = 'TOTAL'
            ws['F31'] = 'TOTAL'
            ws['F38'] = 'TOTAL'
            ws['F47'] = 'TOTAL'
            ws['F53'] = 'TOTAL'
            ws['B2'] = 'Open'
            ws['C2'] = 'Close'
            ws['D2'] = 'Amount'
            ws['B8'] = 'Open'
            ws['C8'] = 'Close'
            ws['D8'] = 'Amount'
            ws['B28'] = 'Open'
            ws['C28'] = 'Close'
            ws['D28'] = 'Amount'
            ws['B41'] = 'Open'
            ws['C41'] = 'Close'
            ws['D41'] = 'Amount'
            ws['G2'] = 'Open'
            ws['H2'] = 'Close'
            ws['I2'] = 'Amount'
            ws['G24'] = 'Open'
            ws['H24'] = 'Close'
            ws['I24'] = 'Amount'
            ws['G34'] = 'Open'
            ws['H34'] = 'Close'
            ws['I34'] = 'Amount'
            ws['G41'] = 'Open'
            ws['H41'] = 'Close'
            ws['I41'] = 'Amount'
            ws['G50'] = 'Open'
            ws['H50'] = 'Close'
            ws['I50'] = 'Amount'

            #BOX NUMBERS
            ws['A3'] = 1
            ws['A4'] = 2
            ws['A9'] = 3
            ws['A10'] = 4
            ws['A11'] = 5
            ws['A12'] = 6
            ws['A13'] = 7
            ws['A14'] = 8
            ws['A15'] = 9
            ws['A16'] = 10
            ws['A17'] = 14
            ws['A18'] = 15
            ws['A19'] = 16
            ws['A20'] = 17
            ws['A21'] = 18
            ws['A22'] = 19
            ws['A23'] = 20
            ws['A24'] = 21
            ws['A29'] = 22
            ws['A30'] = 23
            ws['A31'] = 25
            ws['A32'] = 26
            ws['A33'] = 27
            ws['A34'] = 28
            ws['A35'] = 29
            ws['A36'] = 30
            ws['A37'] = 31
            ws['A42'] = 32
            ws['A43'] = 33
            ws['A44'] = 34
            ws['A45'] = 35
            ws['A46'] = 36
            ws['A47'] = 37
            ws['A48'] = 38
            ws['A49'] = 39
            ws['A50'] = 40
            ws['A51'] = 41
            ws['A52'] = 42
            ws['A53'] = 43
            ws['A54'] = 44
            ws['A55'] = 45
            ws['A56'] = 46

            #Column 2
            ws['F3'] = 47
            ws['F4'] = 48
            ws['F5'] = 49
            ws['F6'] = 50
            ws['F7'] = 51
            ws['F8'] = 52
            ws['F9'] = 53
            ws['F10'] = 54
            ws['F11'] = 55
            ws['F12'] = 56
            ws['F13'] = 57
            ws['F14'] = 58
            ws['F15'] = 59
            ws['F16'] = 60
            ws['F17'] = 61
            ws['F18'] = 62
            ws['F19'] = 63
            ws['F20'] = 64
            ws['F25'] = 65
            ws['F26'] = 66
            ws['F27'] = 67
            ws['F28'] = 68
            ws['F29'] = 69
            ws['F30'] = 70
            ws['F35'] = 13
            ws['F36'] = 24
            ws['F37'] = 71
            ws['F42'] = 72
            ws['F43'] = 73
            ws['F44'] = 74
            ws['F45'] = 75
            ws['F46'] = 76
            ws['F51'] = 77
            ws['F52'] = 78
            
      

            #Formulas
            ws['B5'] = '=SUM(B3:B4)'
            ws['C5'] = '=SUM(C3:C4)'
            ws['D5'] = '=SUM(D3:D4)'
            ws['B25'] = '=SUM(B9:B24)'
            ws['C25'] = '=SUM(C9:C24)'
            ws['D25'] = '=SUM(D9:D24)'
            ws['B38'] = '=SUM(B29:B37)'
            ws['C38'] = '=SUM(C29:C37)'
            ws['D38'] = '=SUM(D29:D37)'
            ws['B57'] = '=SUM(B42:B56)'
            ws['C57'] = '=SUM(C42:C56)'
            ws['D57'] = '=SUM(D42:D56)'

            ws['G21'] = '=SUM(G3:G20)'
            ws['H21'] = '=SUM(H3:H20)'
            ws['I21'] = '=SUM(I3:I20)'

            ws['G31'] = '=SUM(G25:G30)'
            ws['H31'] = '=SUM(H25:H30)'
            ws['I31'] = '=SUM(I25:I30)'

            
            ws['G38'] = '=SUM(G35:G37)'
            ws['H38'] = '=SUM(H35:H37)'
            ws['I38'] = '=SUM(I35:I37)'

            
            ws['G47'] = '=SUM(G42:G46)'
            ws['H47'] = '=SUM(H42:H46)'
            ws['I47'] = '=SUM(I42:I46)'

            
            ws['G53'] = '=SUM(G51:G52)'
            ws['H53'] = '=SUM(H51:H52)'
            ws['I53'] = '=SUM(I51:I52)'

            ws['L24'] = '=M13+M15-M18-M17'

            ws['M4'] = '=D5'
            ws['M5'] = '=D25'
            ws['M6'] = '=D38'
            ws['M7'] = '=D57'
            ws['M8'] = '=I21'
            ws['M9'] = '=I31'
            ws['M10'] = '=I38'
            ws['M11'] = '=I47'
            ws['M12'] = '=I53'

            ws['K3'] = 'Tickets'
            ws['L3'] = 'Sold'
            ws['M3'] = 'Amount'
            ws['K13'] = 'SCRATCH-OFF'
            ws['L13'] = '=SUM(L4:L12)'
            ws['M13'] = '=SUM(M4:M12)'

            ws['K21'] = 'ONLY TYPE IN PURPLE CELLS'
            ws['K15'] = 'ONLINE SALES'
            ws['K17'] = 'ONLINE CASHES'
            ws['K18'] = 'STORE CASHES'
            ws['K24'] = 'TOTAL CASH'
            ws['K26'] = 'TOTAL CASH COUNTED'

            ws['D3'] = '=if(C3<B3,(299-B3+1)+C3,C3-B3)*$A$2'
            ws['D4'] = '=if(C4<B4,(299-B4+1)+C4,C4-B4)*$A$2'
            ws['I51'] = '=if(H51<G51,(17-G51+1)+H51,H51-G51)*$F$50'
            ws['I52'] = '=if(H52<G52,(17-G52+1)+H52,H52-G52)*$F$50'

            def apply_if_formula(ws, start_row, end_row):
            
                for row in range(start_row, end_row + 1):
                    ws[f'D{row}'] = f'=IF(C{row}<B{row},(149-B{row}+1)+C{row},C{row}-B{row})*2'
                    
            # Example usage:
            apply_if_formula(ws, 9, 24)

            def apply_custom_formula(ws, start_row, end_row):

                for row in range(start_row, end_row + 1):
                    ws[f'D{row}'] = f'=IF(C{row}<B{row},(99-B{row}+1)+C{row},C{row}-B{row})*$A$28'

            # Example usage:
            apply_custom_formula(ws, 29, 37)

            def apply_another_formula(ws, start_row, end_row):
 
                for row in range(start_row, end_row + 1):
                    ws[f'D{row}'] = f'=IF(C{row}<B{row},(59-B{row}+1)+C{row},C{row}-B{row})*$A$41'

            # Example usage:
            apply_another_formula(ws, 42, 56)

            def apply_formula_for_I_column(ws, start_row, end_row):
                for row in range(start_row, end_row + 1):
                    ws[f'I{row}'] = f'=IF(H{row}<G{row},(29-G{row}+1)+H{row},H{row}-G{row})*$F$2'

            # Example usage:
            apply_formula_for_I_column(ws, 3, 20)


            def apply_formula_I25_to_I30(ws, start_row, end_row):

                for row in range(start_row, end_row + 1):
                    ws[f'I{row}'] = f'=IF(H{row}<G{row},(14-G{row}+1)+H{row},H{row}-G{row})*$F$24'

            # Example usage:
            apply_formula_I25_to_I30(ws, 25, 30)

            def apply_formula_I35_to_I37(ws, start_row, end_row):

                for row in range(start_row, end_row + 1):
                    ws[f'I{row}'] = f'=IF(H{row}<G{row},(11-G{row}+1)+H{row},H{row}-G{row})*$F$34'

            # Example usage:
            apply_formula_I35_to_I37(ws, 35, 37)

            def apply_formula_I42_to_I46(ws, start_row, end_row):

                for row in range(start_row, end_row + 1):
                    ws[f'I{row}'] = f'=IF(H{row}<G{row},(9-G{row}+1)+H{row},H{row}-G{row})*$F$41'
                    
            # Example usage:
            apply_formula_I42_to_I46(ws, 42, 46)


            def apply_formula(ws, start_row, end_row):
                for row in range(start_row, end_row + 1):
                    ws[f'L{row}'] = f'=M{row}/K{row}'
                    
            apply_formula(ws, 4, 12)  
            

            def fill_range(ws, cell_range, fill_style):
                for row in ws[cell_range]:
                    for cell in row:
                        cell.fill = fill_style

            #Highlight range of cells with color
            
            fill_range(ws, 'C3:C4', lavender_fill)
            fill_range(ws, 'C9:C24', lavender_fill)
            fill_range(ws, 'C29:C37', lavender_fill)
            fill_range(ws, 'C42:C56', lavender_fill)
            fill_range(ws, 'H3:H20', lavender_fill)
            fill_range(ws, 'H25:H30', lavender_fill)
            fill_range(ws, 'H35:H37', lavender_fill)
            fill_range(ws, 'H42:H46', lavender_fill)
            fill_range(ws, 'H51:H52', lavender_fill)
            fill_range(ws, 'K21:KM2', lavender_fill)
            
            #Blue colors for the total
            ws['A5'].fill = blue_fill 
            fill_range(ws, 'B5:D5', LightBlue_fill)

            ws['A25'].fill = blue_fill 
            fill_range(ws, 'B25:D25', LightBlue_fill)

            ws['A38'].fill = blue_fill 
            fill_range(ws, 'B38:D38', LightBlue_fill)

            ws['A57'].fill = blue_fill 
            fill_range(ws, 'B57:D57', LightBlue_fill)

            ws['F21'].fill = blue_fill 
            fill_range(ws, 'G21:I21', LightBlue_fill)

            ws['F31'].fill = blue_fill 
            fill_range(ws, 'G31:I31', LightBlue_fill)

            ws['F38'].fill = blue_fill 
            fill_range(ws, 'G38:I38', LightBlue_fill)

            ws['F47'].fill = blue_fill 
            fill_range(ws, 'G47:I47', LightBlue_fill)
            
            ws['F53'].fill = blue_fill 
            fill_range(ws, 'G53:I53', LightBlue_fill)

            ws['K13'].fill = blue_fill 
            fill_range(ws, 'L13:M13', LightBlue_fill)
            ws['K15'].fill = blue_fill
            ws['K17'].fill = blue_fill
            ws['K18'].fill = blue_fill
            ws['M15'].fill = lavender_fill
            ws['M17'].fill = lavender_fill
            ws['M18'].fill = lavender_fill
            ws['K21'].fill = lavender_fill

            fill_range(ws, 'K24:M24', green_fill)
           
            ws['K26'].fill = green_fill
            ws['L26'].fill = lavender_fill


            # Apply borders to specified cell ranges
            apply_border(ws, 'A2:D5')
            apply_border(ws, 'A8:D25')
            apply_border(ws, 'A28:D38')
            apply_border(ws, 'A41:D57')
            apply_border(ws, 'F2:I21')
            apply_border(ws, 'F24:I31')
            apply_border(ws, 'F34:I38')
            apply_border(ws, 'F41:I47')
            apply_border(ws, 'F50:I53')
            apply_border(ws, 'K2:M13')
            apply_border(ws, 'K15:M15')
            apply_border(ws, 'K17:M18')
            apply_border(ws, 'K21:M21')
            apply_border(ws, 'K24:M24')
            apply_border(ws, 'K26:M26')
            
            


            ws.column_dimensions['K'].width = 23

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')


    # Remove the default sheet created                                                                                                                                                                    
        if 'Sheet' in wb.sheetnames:                                                                                                                                                                            
            wb.remove(wb['Sheet'])     

        link_sheets(wb, sheet_names)                                                                                                                                                                                                      
        # Save the workbook to an Excel file                                                                                                                                                                    
        file_name = f'Generated_Excel_File_{month_number}.xlsx'                                                                                                                                                 
        wb.save(file_name)                                                                                                                                                                                      
                                                                                                                                                                                                                
        print(f'Excel file "{file_name}" has been created successfully.')                                                                                                                                       
                                                                                                                                                                                                                
    else:                                                                                                                                                                                                       
        print("Please enter a valid month number between 1 and 12.")                                                                                                                                            
                                                                                                                                                                                                                
except ValueError as e:                                                                                                                                                                                         
    print("Invalid input. Please enter a numeric value for the month.")                                                                                                                                         
    print("Error details:", e)                                                                                                                                                                                  
    traceback.print_exc()
       
       
    
